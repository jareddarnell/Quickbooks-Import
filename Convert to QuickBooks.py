import openpyxl
import os
import sys
import fnmatch
import datetime
import calendar

# Constants
WorkSheetTitle = "Embroidery"
Customer = "WORLD THREADS, INC"
Item = "EMBROIDERY ITEM"
Price = 8.25
Quantity = 1

# List for headers
headerList = ["Customer", "Transaction Date", "RefNumber", "PO Number",
              "Class", "Template Name", "To Be Printed", "Ship Date",
              "BillTo Line1", "BillTo Line2", "BillTo Line3", "BillTo Line4",
              "BillTo City", "BillTo State", "BillTo PostalCode",
              "BillTo Country", "ShipTo Line2", "ShipTo Line3",
              "ShipTo Line4", "ShipTo City", "ShipTo State",
              "ShipTo PostalCode", "ShipTo Country", "Phone",
              "Fax", "Email", "Contact Name", "First Name",
              "Last Name", "Rep", "Due Date", "Ship Method",
              "Customer Message", "Memo", "Cust. Tax Code",
              "Item", "Quantity", "Description", "Price",
              "Item Line Class", "Service Date", "FOB",
              "Customer Acct No", "Terms", "Sales Tax Item",
              "To Be E-Mailed", "Other", "Other1", "Other2",
              "Unit of Measure", "Currency", "Exchange Rate",
              "Sales Tax Code"]

# List holding lists for storing data before transfering to each new sheet
listofLists = []

# Creates a new sheet with matching headers
def CreateNewSheet(excelFile):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WorkSheetTitle + ' - ' + calendar.month_name[int(excelFile)]
    # Add headers to the first row from the headerList
    for col, val in enumerate(headerList, start=1):
        ws.cell(row=1, column=col).value = val
    
    count = 2
    # Assign item to a cell and then fill cell with data from list
    for list in listofLists:
            customerData = ws.cell(row=count, column=1)
            transactionDateData = ws.cell(row=count, column=2)
            refNumberData = ws.cell(row=count, column=3)
            poNumberData = ws.cell(row=count, column=4)
            shipDateData = ws.cell(row=count, column=8)
            dueDateData = ws.cell(row=count, column=31)
            itemData = ws.cell(row=count, column=36)
            quantityData = ws.cell(row=count, column=37)
            descriptionData = ws.cell(row=count, column=38)
            priceData = ws.cell(row=count, column=39)
            
            customerData.value = list[0]
            transactionDateData.value = list[1]
            refNumberData.value = list[2]
            poNumberData.value = list[3]
            shipDateData.value = list[4]
            dueDateData.value = list[5]
            itemData.value = list[6]
            quantityData.value = list[7]
            descriptionData.value = list[8]
            priceData.value = list[9]
            
            count += 1
        
    wb.save(ws.title + ".xlsx")

# Get current directory
currentDirectory = os.getcwd()

# Change to current working directory
os.chdir(currentDirectory)

# Get list of excel files
excelFiles = fnmatch.filter(os.listdir('.'), '*.xlsx')

# Loop through each Excel file
for i in range(0, len(excelFiles)):

    # Open Excel file
    workBook = openpyxl.load_workbook(excelFiles[i])
    workSheet = workBook.active

    # Create dictionary of column names
    ColNames = {}
    Current  = 0
    for col in workSheet.iter_cols(1, workSheet.max_column):
        ColNames[col[0].value] = col[0].column_letter
        Current += 1
    
    # Get First RefNumber
    if i == 0:
        RefNumber = str(workSheet[ColNames['Order Number'] + "2"].value)[:11]

    # Extract data from sheet and store for later
    for row in workSheet.iter_rows(min_row=2):
        tempList = []
        tempList.append(Customer) # Customer
        
        for cell in row:
            if cell.row == 2:
                if cell.column_letter == ColNames['940 Date/Time Stamp']: # CHANGE TO COLUMN WITH 940 Date/Time Stamp
                    tempList.append(cell.value) # Transaction Date
                if cell.column_letter == ColNames['Order Number']: # CHANGE TO COLUMN WITH Order Number
                    tempList.append(RefNumber) # RefNumber, truncate to 11 characters
                    tempList.append(RefNumber) # PO Number, truncate to 11 characters
                if cell.column_letter == ColNames['Delivery Date']: # CHANGE TO COLUMN WITH Delivery Date
                    tempList.insert(4, cell.value) # Ship Date
                    tempList.insert(5, cell.value) # Due Date
                    tempList.insert(6, Item) # Item
                    tempList.insert(7, (workSheet.max_row - 1)) # Quantity
                if cell.column_letter == ColNames['Mono_Font and Text']: # CHANGE TO COLUMN WITH Mono_Font and Text
                    tempList.insert(8, os.path.splitext(excelFiles[i])[0]) # Name of excel file
                    tempList.insert(9, Price) # Price
        if len(tempList) > 1:
            listofLists.append(tempList)
    
    # Close the Excel file without saving
    workBook.close()

# Create new sheets
CreateNewSheet(str(excelFiles[0])[0:2])

# Exit program
sys.exit()